import scrapy
from openpyxl import load_workbook

wb = load_workbook('BD_CADASTRO_NUMERADO_AGO_TESTE.xlsx')
ws, ws2 = wb['Fundos'], wb['Fundos_Cota']
lista_cnpj = []

for row in ws.iter_rows(values_only=True):
    cnpj = row[1]
    lista_cnpj.append(cnpj)

for row in ws2.iter_rows(values_only=True):
    cnpj = row[1] 
    lista_cnpj.append(cnpj)

class CvmSpider(scrapy.Spider):
    name = "cvm"
    allowed_domains = ["cvmweb.cvm.gov.br"]
    start_urls = ["https://cvmweb.cvm.gov.br/SWB//Sistemas/SCW/CPublica/CConsolFdo/FormBuscaParticFdo.aspx"]

    def parse(self, response):
        for cnpj in lista_cnpj:
            yield scrapy.FormRequest.from_response(
                response,
                formdata={"txtCNPJNome": cnpj},
                callback=self.after_search,
                meta={"cnpj": cnpj}
            )
    
    def after_search(self, response):
        nome_fundo = response.xpath("//a[@id='ddlFundos__ctl0_lnkbtn1']/text()").get()
        event_target = response.xpath("//a[@id='ddlFundos__ctl0_lnkbtn1']/@href").re_first(r"__doPostBack\('([^']+)'")

        if event_target:
            viewstate = response.xpath("//input[@name='__VIEWSTATE']/@value").get()
            eventvalidation = response.xpath("//input[@name='__EVENTVALIDATION']/@value").get()

            yield scrapy.FormRequest(
                url=response.url,
                formdata={
                    '__EVENTTARGET': event_target,
                    '__EVENTARGUMENT': '',
                    '__VIEWSTATE': viewstate,
                    '__EVENTVALIDATION': eventvalidation
                },
                callback=self.parse_fundo,
                meta={"cnpj": response.meta["cnpj"], "nome_fundo": nome_fundo}
            )
    
    def parse_fundo(self, response):
        pagina_fundo = response.url
        administrador_fundo = response.xpath("//span[@id='lbNmDenomSocialAdm']/text()").get()
        situacao_fundo = response.xpath("//span[@id='lbSitDesc']/text()").get()
        inicio_atividades_fundo = response.xpath("//span[@id='lbInfAdc1']/text()").get()
        link = response.xpath("//a[@id='Hyperlink2']/@href").get()
        yield response.follow(url=link, callback=self.parse_dados_diarios, meta={"cnpj": response.meta["cnpj"], "nome_fundo": response.meta["nome_fundo"], 
                                                                                 "pagina_fundo": pagina_fundo, "administrador_fundo": administrador_fundo, 
                                                                                 "situacao_fundo": situacao_fundo, "inicio_atividades_fundo": inicio_atividades_fundo}) 
    
    def parse_dados_diarios(self, response):

        link_voltar = response.meta['pagina_fundo']
        linhas = response.xpath("//table[@id='dgDocDiario']//tr[position()>1]") #ignora a primeira linha, >1 (pula o primeiro tr)

        #se não tiver valor na quota (coluna 2), já filtra e retira a linha
        linhas_validas = [linha for linha in linhas if linha.xpath("td[2]/text()").get()] 

        if linhas_validas:
            ultima_linha = linhas_validas[-1]
            dados = ultima_linha.xpath("td/text()").getall()
            dados_diarios = {
                "Dia": dados[0].strip(),
                "Quota": dados[1].strip(),
                #"Captação no Dia": dados[2].strip(),
                #"Resgate no Dia": dados[3].strip(),
                "Patrimônio Líquido": dados[4].strip(),
                #"Total da Carteira": dados[5].strip(),
                "Número de Cotistas": dados[6].strip(),
                #"Data da Próxima Informação do PL": dados[7].strip(),
            }
        yield response.follow(url=link_voltar, callback=self.parse_clicar_lamina_fundos, meta={"cnpj": response.meta["cnpj"], "nome_fundo": response.meta["nome_fundo"], 
                                                                                 "pagina_fundo": response.meta['pagina_fundo'], "administrador_fundo": response.meta['administrador_fundo'], 
                                                                                 "situacao_fundo": response.meta['situacao_fundo'], "inicio_atividades_fundo": response.meta['inicio_atividades_fundo'],
                                                                                 "dados_diarios": dados_diarios}, dont_filter=True) #o dont_filter serve para override do filtro padrão do scrapy de proibir url repetida.
        
    def parse_clicar_lamina_fundos(self, response):
        link = response.xpath("//a[@id='hlInfLamina']/@href").get()
        yield response.follow(url=link, callback=self.parse_lamina_fundos, meta={"cnpj": response.meta["cnpj"], "nome_fundo": response.meta["nome_fundo"], 
                                                                                               "pagina_fundo": response.meta["pagina_fundo"], "administrador_fundo": response.meta["administrador_fundo"], 
                                                                                               "situacao_fundo": response.meta["situacao_fundo"], "inicio_atividades_fundo": response.meta["inicio_atividades_fundo"],
                                                                                               "dados_diarios": response.meta['dados_diarios']})
    
    def parse_lamina_fundos(self, response):
        print(response.url)
        mes_competencia = response.xpath("//table[@id='Table1']/tbody/tr/td/div/p[4]/b/span").get()
        print("COMPETENCIA??",mes_competencia)
        yield {
            "cnpj": response.meta["cnpj"],
            "nome_fundo": response.meta["nome_fundo"],
            "pagina_fundo": response.meta["pagina_fundo"],
            "administrador_fundo": response.meta["administrador_fundo"],
            "situacao_fundo": response.meta["situacao_fundo"],
            "inicio_atividades_fundo": response.meta["inicio_atividades_fundo"],
            "dados_diarios": response.meta['dados_diarios'],
            "mes_competencia": mes_competencia
        }
